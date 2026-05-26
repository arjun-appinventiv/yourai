"""
Build YourAI_WBS_Five_Modules.xlsx — scope reference for:
  1. YourVault (folder structure + search)
  2. Complete Knowledge Pack
  3. Side Panel of intent outputs (IntentArtifactPanel)
  4. Workflows — Super Admin
  5. Workflows — Chat (Tenant)

Format spec: .claude-context/wbs-format.md
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─── Style constants ─────────────────────────────────────────────
ARIAL = "Arial"
THIN = Side(border_style="thin", color="000000")
ALL_BORDERS = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
YELLOW_FILL = PatternFill("solid", start_color="FFFF00", end_color="FFFF00")
DARKBLUE_FILL = PatternFill("solid", start_color="1F3864", end_color="1F3864")
LIGHTGRAY_FILL = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
TITLE_FONT = Font(name=ARIAL, size=14, bold=True, color="FFFFFF")
HEADER_FONT = Font(name=ARIAL, size=11, bold=True, color="000000")
MODULE_FONT = Font(name=ARIAL, size=11, bold=True, color="000000")
BODY_FONT = Font(name=ARIAL, size=10, color="000000")
BODY_BOLD = Font(name=ARIAL, size=10, bold=True, color="000000")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENTER_TOP = Alignment(horizontal="center", vertical="top", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
COL_WIDTHS = {"A": 8, "B": 28, "C": 25, "D": 60, "E": 16, "F": 38}
HEADERS = ["Sr. No.", "Modules", "Sub-Modules", "Features / WBS", "Status", "Notes / Remarks"]


# ─── Scope content per sheet ─────────────────────────────────────

YOURVAULT = [
    {
        "module": "Vault landing & hero",
        "submods": [
            {
                "name": "Hero + scope tabs",
                "features": [
                    "'Ask across your documents' hero with sparkle gold-bordered NL search bar",
                    "⌘K keyboard shortcut to focus the search input",
                    "Org-Admin scope tabs above hero (All / Org-wide / Mine)",
                    "Upload + New Document buttons top-right of hero",
                    "Folder name + breadcrumb display when inside a folder",
                ],
                "status": "Complete",
                "notes": "Per Aashna 2026-04-27 redesign. Hero stays static; only breadcrumb swaps when navigating into a folder.",
            },
            {
                "name": "Step-numbered toolbar",
                "features": [
                    "1. Visualize — List / Grid toggle (visual only in v1)",
                    "2. {N} docs context counter",
                    "3. Refine — collapses Date / Uploader / Type / Sort chips behind a dropdown",
                    "4. Move — 'Drag a doc into chat' hint",
                    "Chips render only when openFilterMenu === 'refine'",
                ],
                "status": "Complete",
                "notes": "Refine chip overflow added 2026-04-27 to reduce visual noise.",
            },
        ],
    },
    {
        "module": "Folder structure",
        "submods": [
            {
                "name": "Nested folders",
                "features": [
                    "Folders can contain sub-folders to arbitrary depth (v1 ships 2-level mocks)",
                    "Left rail folder list with click-to-navigate",
                    "Folder rows: name + doc-count badge",
                    "MATTERS-only left-rail framing (2026-05-11 PM call dropped user-folders block)",
                    "Cross-matter filter pills under search bar",
                    "Breadcrumb above hero when inside a folder",
                ],
                "status": "Complete",
                "notes": "Folder state in 'yourai_document_vault_folders_v1' localStorage via documentVaultStore.",
            },
            {
                "name": "Folder operations",
                "features": [
                    "Create folder via modal (name + parentId)",
                    "Rename folder via kebab",
                    "Delete folder (deletes contained docs after confirm)",
                    "Move docs between folders (drag or kebab → Move)",
                    "Recursive upload — drag a folder, all contents go to active folder",
                ],
                "status": "Complete",
                "notes": "Wendy P3 / P4 / P5 — nested + recursive + UI label cleanup.",
            },
        ],
    },
    {
        "module": "Document table",
        "submods": [
            {
                "name": "Root view columns",
                "features": [
                    "Folder → Name → Owner → Size → Modified → Actions",
                    "Folder column bolded as primary case-file identifier",
                    "Sort by column header click (Name / Modified)",
                    "Inside-folder view drops the Folder column",
                ],
                "status": "Complete",
                "notes": "Ryan Robertson framing — 'folder' reads as the matter for an attorney's brain.",
            },
            {
                "name": "Per-row actions",
                "features": [
                    "Kebab menu: Open, Download, Rename, Move, Delete",
                    "Click row → opens document detail (right rail — deferred)",
                    "'Use in chat' button when picker is open",
                    "Drag row → attach to chat",
                ],
                "status": "In Progress",
                "notes": "Inspect right-rail panel from designer mockup deferred; click-to-open currently does nothing in root view.",
            },
        ],
    },
    {
        "module": "Search & filter (P8)",
        "submods": [
            {
                "name": "Ask-anything NL parser",
                "features": [
                    "Top search bar with sparkle icon",
                    "NL parsing: 'find indemnification across Acme contracts' → set filters + run search",
                    "AI-set transient filters (resultLimit / askExplanation / sort) drop on user-typed query",
                    "Explicit chip filters (date / uploader / type) persist across typing",
                    "'All documents' tree click also drops transient filters",
                ],
                "status": "Complete",
                "notes": "P8 v1 — Option 2 shipped 2026-04-27.",
            },
            {
                "name": "Filter chips",
                "features": [
                    "Date chip — Last 7d / 30d / This year / All",
                    "Uploader chip — pick from active org users",
                    "Type chip — PDF / DOCX / XLSX / TXT",
                    "Sort chip — Newest / Oldest / Name asc / Name desc",
                    "Clear all link when any chip is active",
                ],
                "status": "Complete",
                "notes": "Chips collapse behind 'Refine' dropdown to keep toolbar compact.",
            },
            {
                "name": "Content-aware search",
                "features": [
                    "Search hits both filename and extracted `content` field",
                    "Highlight matched terms in result row preview",
                    "Empty-state when zero matches: 'No documents match \"X\"'",
                ],
                "status": "Complete",
                "notes": "Substring search over name + description + content; case-insensitive.",
            },
        ],
    },
    {
        "module": "Upload & creation",
        "submods": [
            {
                "name": "Upload pipeline",
                "features": [
                    "Single-file upload via Upload button",
                    "Drag-and-drop into page (folder-aware)",
                    "Real text extraction via extractFileText helper (PDF / DOCX / TXT)",
                    "Per-file processing status (uploading → extracting → ready)",
                    "Failed extraction → 'Flagged for Review' status",
                ],
                "status": "Complete",
                "notes": "Shared src/lib/file-parser.ts pipeline used by chat attach + KP + workflow Pre-Run.",
            },
            {
                "name": "New Document modal",
                "features": [
                    "Field order: Folder → Document name → Description → File",
                    "Folder dropdown defaults to current folder context",
                    "Description field optional",
                    "File picker accepts PDF / DOCX / TXT (≤25 MB)",
                    "Save creates the vault entry + extracts content + closes",
                ],
                "status": "Complete",
                "notes": "Field order changed 2026-04-29 per Ryan Robertson.",
            },
            {
                "name": "Sample seed docs",
                "features": [
                    "4 real PDFs in public/sample-docs/ (NDA, Sample Case Brief, MSA, Compliance Memo)",
                    "Each entry carries extracted `content` so bot can actually read them",
                    "Storage key bumped to yourai_document_vault_v2",
                    "Seed PDFs generated by /tmp/gen-sample-pdfs.py (fpdf2)",
                ],
                "status": "Complete",
                "notes": "Re-run gen-sample-pdfs.py if sample text in sampleVaultContent.ts changes.",
            },
        ],
    },
    {
        "module": "Chat integration",
        "submods": [
            {
                "name": "Vault picker modal",
                "features": [
                    "Triggered from chat '+' button OR from SearchScopePill YourVault option",
                    "Search input + filtered doc list + folder breadcrumbs",
                    "'Use in chat' CTA per row + 'Open YourVault →' footer link",
                    "Detach row when an active vault doc is already pinned",
                    "Closing without picking keeps prior scope",
                ],
                "status": "Complete",
                "notes": "Mirrors KP picker modal pattern; both use same search-first layout.",
            },
            {
                "name": "Scope-as-conversation-context",
                "features": [
                    "Selected vault doc inlined into messageForEdge under [Documents attached to this conversation]",
                    "Per-doc cap 5,000 chars in the system prompt",
                    "Source pill 'Answered from: <doc name>' on grounded responses",
                ],
                "status": "Complete",
                "notes": "Scope commits only when the user actually picks a doc inside the modal.",
            },
        ],
    },
    {
        "module": "Permissions & visibility",
        "submods": [
            {
                "name": "Role-based scope",
                "features": [
                    "External User → only their own docs / docs shared into their workspace",
                    "Internal / Manager → org-wide + their own",
                    "Org Admin → everything",
                    "Mine vs Org-wide flag per doc",
                ],
                "status": "Complete",
                "notes": "Filter lives inside DocumentVaultPanel based on currentRole.",
            },
            {
                "name": "Sharing & promotion",
                "features": [
                    "Promote 'Mine' doc to 'Org-wide' via kebab",
                    "Confirmation prompt before promoting",
                    "Visibility filter re-evaluated after promote",
                    "Audit-log entry on every promote (planned)",
                ],
                "status": "In Progress",
                "notes": "Promote path works; audit-log entry needs the audit pipeline to be wired.",
            },
        ],
    },
    {
        "module": "Backend-pending (Sprint 2+)",
        "submods": [
            {
                "name": "Metadata enrichment (P8.3)",
                "features": [
                    "Page count via best-effort PDF parse",
                    "lastModifiedAt stamp",
                    "Auto-stamp workspaceId on workspace-attached docs",
                    "Enables 'files over 200 pages' / 'files in Acme matter' queries",
                ],
                "status": "Phase 2/3",
                "notes": "Wendy P8.3 — backend-dependent. Without metadata, certain NL queries can't return useful results.",
            },
            {
                "name": "Content semantic search (P8.4)",
                "features": [
                    "pgvector ingestion pipeline",
                    "Hybrid retrieval (dense + BM25 + RRF + rerank)",
                    "Per-file token math (chunking, overlap, dedup)",
                    "Async ingestion on upload",
                    "~3-week effort per .claude-context/vault-content-rag-plan.md",
                ],
                "status": "Phase 2/3",
                "notes": "Wendy P8.4 — the real grounding engine. Today's substring search is a stopgap.",
            },
            {
                "name": "Large-file strategy (P12)",
                "features": [
                    "Auto-chunking for 500-page PDFs (family-law app-close exports)",
                    "Per-case MD summary index",
                    "Possible open-source memory system per Hog",
                ],
                "status": "Phase 2/3",
                "notes": "Wendy P12 — backend-dependent.",
            },
        ],
    },
]


KNOWLEDGE_PACK = [
    {
        "module": "Pack catalog (landing page)",
        "submods": [
            {
                "name": "Card-grid layout",
                "features": [
                    "Vertical pack cards (post-2026-04-27 Aashna redesign — flipped from horizontal)",
                    "Hero with 'Knowledge Packs' title + Create New Pack button",
                    "Filter tabs: All / Org-wide / Mine",
                    "Search bar (filters by pack name + description)",
                    "Per-card doc count + link count + scope badge",
                ],
                "status": "Complete",
                "notes": "Same full-page chrome as Vault / Workspaces / Workflows.",
            },
            {
                "name": "Pack actions per card",
                "features": [
                    "Click card → opens Edit modal",
                    "Kebab: Rename / Duplicate / Delete / Share toggle",
                    "Active-pack indicator (gold pill) when pack is currently attached to chat",
                ],
                "status": "Complete",
                "notes": "Active-pack state lives on ChatView, lifted through props.",
            },
        ],
    },
    {
        "module": "Pack lifecycle (create / edit / delete)",
        "submods": [
            {
                "name": "EditKnowledgePackModal",
                "features": [
                    "Pack name (required) + description (optional)",
                    "Scope toggle: Mine / Org-wide",
                    "Add documents tab — file picker (PDF / DOCX / TXT)",
                    "Add links tab — URL + label (metadata-only in v1)",
                    "Per-doc preview row with size + extracted-content status",
                    "Save / Cancel buttons (Save = green when valid)",
                ],
                "status": "Complete",
                "notes": "Single modal handles both Create and Edit flows via existing-pack-id prop.",
            },
            {
                "name": "Document extraction",
                "features": [
                    "runDocExtraction() calls extractFileText (same parser as chat attach)",
                    "Real PDF / DOCX / TXT parsing on upload",
                    "Status flow: uploading → extracting → ready",
                    "Per-doc cap 5,000 chars when inlined into chat",
                    "Strips File ref before save; persists name / size / uploaded / content",
                ],
                "status": "Complete",
                "notes": "Replaced 2026-05-06 simulateDocPipeline stub with real extraction; v1→v2 storage bump forced re-seed.",
            },
            {
                "name": "Pack deletion",
                "features": [
                    "Delete confirmation modal",
                    "Disable delete when pack is currently active in chat",
                    "Toast confirmation post-delete",
                ],
                "status": "Complete",
                "notes": "",
            },
        ],
    },
    {
        "module": "Seed packs",
        "submods": [
            {
                "name": "Bundled defaults",
                "features": [
                    "Hartwell NDA Template Pack — NDA template + risk checklist + redline",
                    "M&A Diligence Pack — checklist + Meridian precedent + indemnification clauses",
                    "California Labor Law Pack — CA Labor Code summary + § 16600 / Edwards rule",
                    "Privacy Pack — GDPR + CCPA reference",
                    "All carry real extracted `content` from src/data/samplePackContent.ts",
                ],
                "status": "Complete",
                "notes": "Seeded via seedPacksIfEmpty(DEFAULT_KNOWLEDGE_PACKS) on first mount.",
            },
        ],
    },
    {
        "module": "Pack picker (chat integration)",
        "submods": [
            {
                "name": "Modal picker",
                "features": [
                    "Triggered from Pack pill in chat input row",
                    "Search input (sticky when > 5 packs)",
                    "Filtered pack list with 'Use' CTA per row",
                    "'No pack' row at top to clear active pack",
                    "'Manage knowledge packs →' footer link",
                ],
                "status": "Complete",
                "notes": "Mirrors YourVault doc-picker modal for consistency.",
            },
            {
                "name": "Pack pill in chat input",
                "features": [
                    "Shows active pack name when one is attached",
                    "'⌘ Knowledge pack' label when no pack active",
                    "Click opens picker modal",
                    "Gold styling when active (matches 'Using {pack} pack' status line)",
                ],
                "status": "Complete",
                "notes": "Pack pill is independent of Intent and SearchScope pills.",
            },
        ],
    },
    {
        "module": "Grounding & chat prompt injection",
        "submods": [
            {
                "name": "Edge prompt payload",
                "features": [
                    "Active pack content prepended to messageForEdge under [Knowledge Pack reference for this conversation]",
                    "Per-doc cap 5,000 chars (truncated with notice)",
                    "Pack content sent alongside attached doc content (both inlined)",
                    "Source pill 'Answered from Pack: {pack name}' rendered when pack-grounded",
                ],
                "status": "Complete",
                "notes": "Edge route (api/chat.ts) is where the inline happens — client fallback (callLLM) does NOT do this.",
            },
            {
                "name": "Intent orthogonality",
                "features": [
                    "Changing intent never changes active pack",
                    "Changing pack never changes active intent",
                    "Independent session-level states (per Conventions doc)",
                ],
                "status": "Complete",
                "notes": "By design — attorney can pair any intent with any pack.",
            },
        ],
    },
    {
        "module": "Permissions & sharing",
        "submods": [
            {
                "name": "Visibility filter",
                "features": [
                    "Mine packs visible to owner only",
                    "Org-wide packs visible to every member of the org",
                    "External User → no pack picker (KP feature gated off externals)",
                ],
                "status": "Complete",
                "notes": "isExternalUser check inside Sidebar gates the Knowledge packs nav item.",
            },
            {
                "name": "Share / promote toggle",
                "features": [
                    "Toggle in EditKnowledgePackModal: Mine ↔ Org-wide",
                    "Confirmation prompt when promoting (irreversible without an admin)",
                    "Visibility re-evaluates immediately",
                ],
                "status": "In Progress",
                "notes": "UI exists; visibility filter logic should be re-verified once a second user account is in localStorage (per 2026-05-06 session note).",
            },
        ],
    },
    {
        "module": "Persistence",
        "submods": [
            {
                "name": "localStorage store",
                "features": [
                    "Key: yourai_knowledge_packs_v2",
                    "Mirrors documentVaultStore pattern (loadPacks / savePacks / seedPacksIfEmpty)",
                    "v2 bump 2026-05-07 forced re-seed after content-field fix",
                    "Save on every mutation via useEffect",
                ],
                "status": "Complete",
                "notes": "src/lib/knowledgePackStore.ts.",
            },
        ],
    },
    {
        "module": "Backend-pending (Sprint 2+)",
        "submods": [
            {
                "name": "Link fetching pipeline",
                "features": [
                    "Save link with URL + title + extracted body",
                    "CORS + sanitization layer for arbitrary URLs",
                    "Re-fetch on link rot (broken link badge)",
                    "Inline link content into chat prompt (like doc content)",
                ],
                "status": "Phase 2/3",
                "notes": "Links save metadata-only today; no grounding. Sprint-2 lift due to CORS + link-rot complexity.",
            },
            {
                "name": "Server-side pack dedup",
                "features": [
                    "Detect duplicate doc content across packs",
                    "Suggest 'use existing doc from Pack X' on upload",
                    "Reduce token cost when multiple packs share a template",
                ],
                "status": "Phase 2/3",
                "notes": "Backend-dependent.",
            },
            {
                "name": "Pack analytics",
                "features": [
                    "Which packs are most often attached to chat",
                    "Which docs in a pack get cited most",
                    "Per-attorney pack usage in Usage & Costs panel",
                ],
                "status": "Phase 2/3",
                "notes": "Requires audit log + cost telemetry.",
            },
        ],
    },
]


SIDE_PANEL = [
    {
        "module": "Inline preview chip",
        "submods": [
            {
                "name": "Chat-thread chip",
                "features": [
                    "Eyebrow line (intent name uppercase — RISK MEMO / SUMMARY / etc)",
                    "DM Serif title (matter / document / 'N findings')",
                    "'Open' pill when panel is closed; 'Viewing' pill when open and showing this artifact",
                    "Click → opens panel anchored to this message",
                    "Per-intent accent color on the chip border",
                ],
                "status": "Complete",
                "notes": "Replaces inline IntentCard render (which was blowing out the chat flow).",
            },
        ],
    },
    {
        "module": "Right-rail panel layout",
        "submods": [
            {
                "name": "Docked panel",
                "features": [
                    "540 px width sibling of chat-main (chat shrinks; NOT an overlay)",
                    "White background (flipped from warm cream 2026-05-04)",
                    "Auto-opens on first card-intent arrival",
                    "Anchored to message ID via activeArtifactMsgId",
                    "Header: Copy as Markdown / Fullscreen / Close buttons",
                ],
                "status": "Complete",
                "notes": "Sibling-panel pattern same as MyTimePanel / AuditLogsPanel.",
            },
            {
                "name": "Fullscreen mode",
                "features": [
                    "Centered ~720 px column",
                    "Same header controls (Copy / Exit fullscreen / Close)",
                    "Exit fullscreen returns to docked mode",
                    "Body remains scrollable",
                ],
                "status": "Complete",
                "notes": "",
            },
            {
                "name": "Markdown body",
                "features": [
                    "ReactMarkdown rendering with 720 px column",
                    "h1 + meta line + h2 sections + bullets + plain blockquotes",
                    "DM Sans body, DM Serif only for h1, h2 with subtle border-bottom",
                    "No pills / tiles / coloured rails — pure report layout",
                ],
                "status": "Complete",
                "notes": "PM call: 'should look like a report, not fancy shit.'",
            },
        ],
    },
    {
        "module": "Card → markdown serializers",
        "submods": [
            {
                "name": "Per-intent converters",
                "features": [
                    "src/lib/cardToMarkdown.ts — 7 serializers (Summary, Comparison, CaseBrief, Research, RiskMemo, ClauseAnalysis, Timeline)",
                    "Each serializer turns structured card JSON into clean markdown",
                    "pickTitle() heuristic — rejects 'Risk Assessment of Uploaded Documents' / 'Untitled case' / 'Legal Inquiry'",
                    "Generic-title regex auto-falls-back to documentName (extension stripped)",
                    "Centralised so adding a new intent = add a serializer; panel picks it up automatically",
                ],
                "status": "Complete",
                "notes": "Timeline serializer remains but Timeline intent itself was retired.",
            },
            {
                "name": "Empty-state branch",
                "features": [
                    "Per-card detection (schema-shaped envelope w/ no real data)",
                    "Empty-state card variant renders in panel instead of grids of '—'",
                    "Upload prompt + sibling-intent hint inside the empty state",
                    "find_document keeps inline FileResultsCard — does NOT use the panel",
                ],
                "status": "Complete",
                "notes": "Detection rule: all schema-required string fields blank + no documentName + empty arrays. See .claude-context/card-empty-state-pattern.md.",
            },
        ],
    },
    {
        "module": "State management",
        "submods": [
            {
                "name": "Panel state",
                "features": [
                    "activeArtifactMsgId on ChatView — anchors the panel to a specific bot message",
                    "Opens automatically on new card-intent arrival",
                    "Switching demos (/demo-risk → /demo-summary) updates panel content",
                    "Close button → setActiveArtifactMsgId(null)",
                    "Click chip after close → re-opens panel to that artifact",
                ],
                "status": "Complete",
                "notes": "",
            },
        ],
    },
    {
        "module": "Copy + share",
        "submods": [
            {
                "name": "Copy-as-Markdown",
                "features": [
                    "Copy button in panel header",
                    "Writes plain markdown text to clipboard",
                    "Tooltip reads 'Copy as Markdown'",
                    "Toast confirmation post-copy",
                ],
                "status": "Complete",
                "notes": "Tooltip text fixed 2026-05-04 — was incorrectly reading 'Copy as JSON'.",
            },
            {
                "name": "Share / export",
                "features": [
                    "PDF export of the report",
                    "Share-as-link (URL with artifact ID)",
                    "Citation linking back to attached doc passages",
                    "Print-stylesheet for browser print",
                ],
                "status": "Phase 2/3",
                "notes": "Browser print works today as a stopgap. Real share-as-link needs backend persistence.",
            },
        ],
    },
    {
        "module": "Backend-pending",
        "submods": [
            {
                "name": "Server-side artifact persistence",
                "features": [
                    "Save generated artifacts to a backend store",
                    "Re-open a prior artifact via thread history",
                    "Versioning (regenerated artifacts keep history)",
                    "Cross-device sync",
                ],
                "status": "Phase 2/3",
                "notes": "Today artifacts live only in chat-thread state; lost on refresh of an old thread.",
            },
        ],
    },
]


SA_WORKFLOWS = [
    {
        "module": "Templates list (SA portal)",
        "submods": [
            {
                "name": "Table view",
                "features": [
                    "List of all workflow templates across all tenants",
                    "Columns: Name / Practice Area / Step Count / Last Edited / Status / Actions",
                    "Search by template name or description",
                    "Filter by practice area + status (Active / Draft / Archived)",
                    "Create New Template button top-right",
                ],
                "status": "Complete",
                "notes": "Lives at /super-admin/workflows. Backed by mockData; switch to real API once dev team ships.",
            },
        ],
    },
    {
        "module": "Template editor",
        "submods": [
            {
                "name": "Step builder",
                "features": [
                    "Centered hero with step-pill indicator",
                    "White rounded panel body",
                    "Navy step-number circles per step",
                    "Drag-to-reorder steps",
                    "Add / remove / duplicate step",
                    "Uppercase mono section labels",
                ],
                "status": "Complete",
                "notes": "Picker + Builder rewritten 2026-04-25 from Aashna's chat-mode mockups.",
            },
            {
                "name": "Per-step config",
                "features": [
                    "Operation type dropdown (Risk / Summary / Comparison / Case Brief / Research / Clause Analysis)",
                    "System prompt textarea (per-step override)",
                    "Output type (card / markdown / table)",
                    "Reference doc per step — warm-beige inset, vault picker",
                    "Optional skip flag (conditional execution)",
                ],
                "status": "Complete",
                "notes": "Prompts default to workflowPrompts.ts operation library.",
            },
            {
                "name": "Save & validation",
                "features": [
                    "Save Draft / Publish buttons",
                    "Required-field validation per step",
                    "Test run inline (executes against a sample doc)",
                    "Auto-save on field blur",
                ],
                "status": "In Progress",
                "notes": "Test-run inline is mocked; needs real execution wiring same as chat workflows.",
            },
        ],
    },
    {
        "module": "Operation library",
        "submods": [
            {
                "name": "Base operations",
                "features": [
                    "6 base ops — Risk Assessment, Document Summarization, Clause Comparison, Case Brief, Legal Research, Clause Analysis",
                    "v2-tightened prompts (word targets + citation hedging + table syntax + findings-vs-actions distinctness)",
                    "All defined in workflowPrompts.ts",
                    "Anti-hallucination string: 'Not covered by supplied documents.' (literal, every operation prompt)",
                ],
                "status": "Complete",
                "notes": "v2 tightening 2026-04-24, blind pass.",
            },
            {
                "name": "Custom operations",
                "features": [
                    "Define a new op from scratch (name + system prompt + output schema)",
                    "Promote a per-step prompt back into the global library",
                    "Versioning per op",
                ],
                "status": "Planned",
                "notes": "Today ops are hard-coded in workflowPrompts.ts; making them editable requires backend storage.",
            },
        ],
    },
    {
        "module": "Distribution",
        "submods": [
            {
                "name": "Tenant visibility",
                "features": [
                    "Active / Inactive flag per template",
                    "Per-tenant override (push to a specific firm)",
                    "Default-all-tenants behavior on publish",
                    "Audit row on every distribution event",
                ],
                "status": "In Progress",
                "notes": "Distribution is mocked end-to-end; backend pipeline needed.",
            },
        ],
    },
    {
        "module": "Audit & versioning",
        "submods": [
            {
                "name": "Edit history",
                "features": [
                    "Last edited by / when displayed on each template",
                    "Full version history viewer (diff per step)",
                    "Rollback to a prior version",
                    "Approve / reject changes (multi-admin flow)",
                ],
                "status": "Planned",
                "notes": "Single-line 'Last edited' shown today; full history viewer is backend-dependent.",
            },
        ],
    },
    {
        "module": "Backend-pending",
        "submods": [
            {
                "name": "Real distribution pipeline",
                "features": [
                    "Push template changes to live tenant chats without redeploy",
                    "Tenant-level enable/disable for in-flight runs",
                    "Versioned rollout (canary 10% → 50% → 100%)",
                ],
                "status": "Phase 2/3",
                "notes": "Backend-dependent.",
            },
            {
                "name": "Usage analytics per template",
                "features": [
                    "Run count per template (today / 7d / 30d)",
                    "Average duration per step",
                    "Failure rate per step + per template",
                    "Surface in SA Reports & Analytics",
                ],
                "status": "Phase 2/3",
                "notes": "Telemetry pipeline needed.",
            },
        ],
    },
]


CHAT_WORKFLOWS = [
    {
        "module": "Workflow picker",
        "submods": [
            {
                "name": "Unified grid",
                "features": [
                    "Single unified grid (no maxWidth cap)",
                    "AI PIPELINES eyebrow row",
                    "Per-card: practice-area top stripe + name + short description",
                    "StatTiles per card (runs in last 7d / avg duration / success rate)",
                    "Running-in pill when a run from that template is active",
                    "PIPELINE op-icon row at card bottom (per-step operation icons)",
                ],
                "status": "Complete",
                "notes": "Rewritten 2026-04-25 from Aashna's chat-mode mockups. The earlier auto-fit + maxWidth 960 cap was retired same day.",
            },
            {
                "name": "Filter & search",
                "features": [
                    "Underline filter tabs (All / Recent / Favourites)",
                    "Search by template name + description",
                    "Click card → opens Pre-Run modal (if doc-required) or runs directly",
                ],
                "status": "Complete",
                "notes": "Favourites are localStorage-only.",
            },
        ],
    },
    {
        "module": "Pre-Run modal",
        "submods": [
            {
                "name": "Document gathering",
                "features": [
                    "Multi-file upload",
                    "Drag-and-drop into modal",
                    "Per-file processing status (extracting / ready / failed)",
                    "classifyDocs() auto-classification on upload-ready",
                    "Per-row uppercase type chip (Contract / Brief / Discovery / Pleading)",
                    "Summary banner ('3 contracts, 2 briefs') above doc list",
                ],
                "status": "Complete",
                "notes": "Classification wired into PreRunModal 2026-04-24 afternoon.",
            },
            {
                "name": "Run kickoff",
                "features": [
                    "Run button enabled when min-doc-count satisfied",
                    "AlreadyRunningAlert guard if another run is active",
                    "Closes modal + auto-opens WorkflowRunPanel",
                ],
                "status": "Complete",
                "notes": "One-run-at-a-time guard prevents concurrent runs.",
            },
        ],
    },
    {
        "module": "Workflow Builder (tenant)",
        "submods": [
            {
                "name": "Step editor",
                "features": [
                    "Centered hero + step-pill indicator",
                    "Add / remove / reorder steps",
                    "Per-step: operation pick + system prompt edit + reference-doc picker",
                    "Save Draft / Save & Run",
                    "In-panel CTAs at bottom",
                ],
                "status": "Complete",
                "notes": "Tenant builder mirrors SA editor with reduced permissions (no template publishing).",
            },
        ],
    },
    {
        "module": "Run execution",
        "submods": [
            {
                "name": "workflowRunner (singleton)",
                "features": [
                    "Module-level singleton with subscribe / notify",
                    "subscribeRun(runId, cb) — components watch one run by id",
                    "Survives component unmount + panel swap (artifact panel pattern)",
                    "getActiveRunId() / getRun(id) / listRuns() exports",
                ],
                "status": "Complete",
                "notes": "src/lib/workflowRunner.ts. Same pattern as sessionTimer (AI-time meter).",
            },
            {
                "name": "workflowExecutor",
                "features": [
                    "Per-step LLM call via /api/chat (Edge function)",
                    "Step N receives summary of steps 1..N-1 (prior-step chaining)",
                    "Per-step prompt assembly: operation system prompt + user instruction + prior outputs (capped ~3500 chars each) + uploaded docs (capped ~8000 chars each) + reference doc + workspace context",
                    "Streams response via ReadableStream",
                    "90s timeout per step via AbortController",
                    "Failure on step → run status 'failed' + error string per step",
                ],
                "status": "Complete",
                "notes": "src/lib/workflowExecutor.ts. The prior-step chaining is the whole point — Generate Report synthesizes prior outputs, not raw docs.",
            },
            {
                "name": "Run states",
                "features": [
                    "queued → running (currentStepIndex 0..N-1) → complete | failed | cancelled",
                    "Cancellation: in-flight AbortController fires",
                    "Per-step timing stored on the run object",
                ],
                "status": "Complete",
                "notes": "",
            },
        ],
    },
    {
        "module": "Run Panel UI",
        "submods": [
            {
                "name": "WorkflowRunPanel",
                "features": [
                    "Multi-run list (active + recent)",
                    "Per-run row: template name + status + currentStep / total",
                    "Click row → expands to step-by-step status",
                    "Cancel button when running",
                    "Re-run button when complete / failed",
                ],
                "status": "Complete",
                "notes": "Sibling panel of chat-main.",
            },
            {
                "name": "WorkflowProgressCard (in-thread)",
                "features": [
                    "Appears in chat thread when a run is triggered from chat",
                    "Pulsing dot + step name + 'View progress →' link",
                    "Updates on subscribeRun ticks",
                    "Collapses to summary line on complete",
                ],
                "status": "Complete",
                "notes": "scrollToRunningPanel() scrolls the chat to the in-thread card on demand.",
            },
        ],
    },
    {
        "module": "Report card",
        "submods": [
            {
                "name": "WorkflowReportCard (Option D)",
                "features": [
                    "Document-style report layout (h1 + meta + per-step sections)",
                    "Per-step output rendered in the section's body",
                    "Final synthesis step as the Conclusion",
                    "Renders inside the chat thread below the progress card",
                    "Citation badges per finding",
                ],
                "status": "Complete",
                "notes": "Option D shipped 2026-04-24. Embedded in chat thread, not in IntentArtifactPanel.",
            },
            {
                "name": "Export from report",
                "features": [
                    "Copy as Markdown",
                    "Print (browser stylesheet)",
                    "Download PDF",
                ],
                "status": "In Progress",
                "notes": "Copy as Markdown works; PDF export is browser-print stopgap.",
            },
        ],
    },
    {
        "module": "Persistence & history",
        "submods": [
            {
                "name": "localStorage runs",
                "features": [
                    "Key: yourai_workflow_runs_v1",
                    "Stores last ~50 runs",
                    "Survives refresh + reopens to last-active run on demand",
                ],
                "status": "Complete",
                "notes": "",
            },
            {
                "name": "Run history view",
                "features": [
                    "Filter by template / status / date range",
                    "Reopen any past run's report card",
                    "Bulk delete / archive",
                ],
                "status": "In Progress",
                "notes": "List exists; filter + reopen wiring needs polish.",
            },
        ],
    },
    {
        "module": "Backend-pending",
        "submods": [
            {
                "name": "Server-side run store",
                "features": [
                    "Move runs out of localStorage into a real backend",
                    "Cross-device sync (start on laptop, view on phone)",
                    "Per-attorney run history surfaces in Reports & Analytics",
                    "Real-time collaboration on a single run",
                ],
                "status": "Phase 2/3",
                "notes": "localStorage is fine for v1 but doesn't scale beyond a single browser.",
            },
            {
                "name": "Cost telemetry per run",
                "features": [
                    "Token cost per step",
                    "Total run cost",
                    "Aggregate per template in Usage & Costs",
                ],
                "status": "Phase 2/3",
                "notes": "",
            },
        ],
    },
]


SHEETS = [
    ("1. YourVault", YOURVAULT),
    ("2. Knowledge Pack", KNOWLEDGE_PACK),
    ("3. Side Panel (Intent Outputs)", SIDE_PANEL),
    ("4. Workflows - Super Admin", SA_WORKFLOWS),
    ("5. Workflows - Chat (Tenant)", CHAT_WORKFLOWS),
]


# ─── Helpers ─────────────────────────────────────────────────────

def status_to_pct(status):
    s = status.lower()
    if "complete" in s and "in progress" not in s: return 100
    if "in progress" in s or "partial" in s: return 50
    if "planned" in s: return 10
    if "phase" in s: return 0
    return 50


def apply_borders(ws, max_row, max_col=6):
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = ALL_BORDERS


def build_platform_sheet(ws, sheet_name, modules):
    # Column widths
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = sheet_name
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = DARKBLUE_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 32

    # Headers
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = YELLOW_FILL
        cell.alignment = CENTER
    ws.row_dimensions[2].height = 30

    # Data
    row = 3
    for mod_idx, module in enumerate(modules, start=1):
        sr_no = f"{mod_idx}.0"
        first_data_row = row
        for sub in module["submods"]:
            features_text = "\n".join(f"- {f}" for f in sub["features"])
            # Sr.No + Module name in the FIRST sub-row only
            if row == first_data_row:
                cell_a = ws.cell(row=row, column=1, value=sr_no)
                cell_a.font = MODULE_FONT
                cell_a.alignment = CENTER

                cell_b = ws.cell(row=row, column=2, value=module["module"])
                cell_b.font = MODULE_FONT
                cell_b.alignment = CENTER

            cell_c = ws.cell(row=row, column=3, value=sub["name"])
            cell_c.font = BODY_FONT
            cell_c.alignment = CENTER_TOP

            cell_d = ws.cell(row=row, column=4, value=features_text)
            cell_d.font = BODY_FONT
            cell_d.alignment = LEFT_TOP

            cell_e = ws.cell(row=row, column=5, value=sub["status"])
            cell_e.font = BODY_FONT
            cell_e.alignment = CENTER_TOP

            cell_f = ws.cell(row=row, column=6, value=sub["notes"])
            cell_f.font = BODY_FONT
            cell_f.alignment = LEFT_TOP

            n = len(sub["features"])
            ws.row_dimensions[row].height = max(40, min(320, n * 20 + 12))

            row += 1

        last_data_row = row - 1
        if last_data_row > first_data_row:
            ws.merge_cells(start_row=first_data_row, start_column=1,
                           end_row=last_data_row, end_column=1)
            ws.merge_cells(start_row=first_data_row, start_column=2,
                           end_row=last_data_row, end_column=2)

    apply_borders(ws, max_row=row - 1)
    ws.freeze_panes = "A3"


def build_overview_sheet(ws, sheet_summaries, grand_pct):
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    ws.merge_cells("A1:E1")
    ws["A1"] = "Overview — Five-Module WBS"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = DARKBLUE_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 32

    headers = ["Module Area", "Modules", "Sub-Modules", "Features", "Completion %"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = YELLOW_FILL
        c.alignment = CENTER
    ws.row_dimensions[2].height = 30

    row = 3
    for s in sheet_summaries:
        ws.cell(row=row, column=1, value=s["name"]).font = BODY_FONT
        ws.cell(row=row, column=1).alignment = LEFT_TOP
        for col_idx, key in enumerate(["modules", "submods", "features"], start=2):
            c = ws.cell(row=row, column=col_idx, value=s[key])
            c.font = BODY_FONT
            c.alignment = CENTER_TOP
        pct_cell = ws.cell(row=row, column=5, value=s["pct"] / 100.0)
        pct_cell.font = BODY_FONT
        pct_cell.alignment = CENTER_TOP
        pct_cell.number_format = "0.0%"
        ws.row_dimensions[row].height = 24
        row += 1

    # Grand total
    total_modules = sum(s["modules"] for s in sheet_summaries)
    total_submods = sum(s["submods"] for s in sheet_summaries)
    total_features = sum(s["features"] for s in sheet_summaries)

    ws.cell(row=row, column=1, value="Grand Total")
    ws.cell(row=row, column=2, value=total_modules)
    ws.cell(row=row, column=3, value=total_submods)
    ws.cell(row=row, column=4, value=total_features)
    pct_cell = ws.cell(row=row, column=5, value=grand_pct / 100.0)
    pct_cell.number_format = "0.0%"

    for col_idx in range(1, 6):
        c = ws.cell(row=row, column=col_idx)
        c.font = BODY_BOLD
        c.fill = LIGHTGRAY_FILL
        c.alignment = CENTER if col_idx > 1 else LEFT_TOP
    ws.cell(row=row, column=1).font = BODY_BOLD
    ws.row_dimensions[row].height = 28
    row += 1

    # Empty separator
    row += 1

    # Status legend
    legend = [
        ("Complete", "Fully implemented with no TODOs (100%)"),
        ("In Progress", "Partially implemented or has TODOs / placeholders (50%)"),
        ("Planned", "Surface scaffolded but not built (10%)"),
        ("Phase 2/3", "Backlog only — depends on backend or future sprint (0%)"),
    ]
    ws.cell(row=row, column=1, value="STATUS LEGEND").font = BODY_BOLD
    ws.cell(row=row, column=1).fill = LIGHTGRAY_FILL
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    ws.cell(row=row, column=2).fill = LIGHTGRAY_FILL
    row += 1
    for label, desc in legend:
        ws.cell(row=row, column=1, value=label).font = BODY_BOLD
        ws.cell(row=row, column=1).alignment = LEFT_TOP
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        c = ws.cell(row=row, column=2, value=desc)
        c.font = BODY_FONT
        c.alignment = LEFT_TOP
        row += 1

    apply_borders(ws, max_row=row - 1, max_col=5)


# ─── Build ───────────────────────────────────────────────────────

wb = Workbook()
# Remove default sheet — we'll create Overview first
wb.remove(wb.active)

# Build platform sheets first to compute summaries
sheet_summaries = []
sheets_built = []
for sheet_name, modules in SHEETS:
    ws = wb.create_sheet(title=sheet_name[:31])
    build_platform_sheet(ws, sheet_name, modules)

    n_modules = len(modules)
    n_submods = sum(len(m["submods"]) for m in modules)
    n_features = sum(len(s["features"]) for m in modules for s in m["submods"])
    avg_pct = sum(status_to_pct(s["status"]) for m in modules for s in m["submods"]) / max(n_submods, 1)
    sheet_summaries.append({
        "name": sheet_name,
        "modules": n_modules,
        "submods": n_submods,
        "features": n_features,
        "pct": avg_pct,
    })

# Grand pct — weighted by sub-module count
total_submods = sum(s["submods"] for s in sheet_summaries)
grand_complete = sum(s["submods"] * (s["pct"] / 100.0) for s in sheet_summaries)
grand_pct = (grand_complete / total_submods) * 100 if total_submods else 0

# Insert Overview at the start
overview = wb.create_sheet(title="Overview", index=0)
build_overview_sheet(overview, sheet_summaries, grand_pct)

OUT_PRIMARY = "/Users/admin/Downloads/scope-creator-ai/.claude/worktrees/great-banach/docs/extracted/YourAI_WBS_Five_Modules.xlsx"
OUT_DESKTOP = "/Users/admin/Desktop/YourAI_WBS_Five_Modules.xlsx"

wb.save(OUT_PRIMARY)
wb.save(OUT_DESKTOP)

print(f"Wrote {OUT_PRIMARY}")
print(f"Wrote {OUT_DESKTOP}")
print()
print("Summary:")
for s in sheet_summaries:
    print(f"  {s['name']:<42} {s['modules']:>3} modules · {s['submods']:>3} sub-modules · {s['features']:>4} features · {s['pct']:>5.1f}%")
print(f"  {'-' * 42}  {'-' * 50}")
print(f"  {'GRAND TOTAL':<42} {sum(s['modules'] for s in sheet_summaries):>3} modules · {total_submods:>3} sub-modules · {sum(s['features'] for s in sheet_summaries):>4} features · {grand_pct:>5.1f}%")
